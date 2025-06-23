import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import win32com.client as win32
import win32process
from contextlib import contextmanager
import psutil
import os

def create_excel_pivot_sumifs(datafile, outputfile='excel_pivot.xlsx', pivot_params=None):
    """
    Create live Excel pivot table from a CSV file, using conditional formulas SUMIFS, COUNTIFS etc.

    Args:
        - datafile (str): Path to CSV data file
        - outputfile (str, optional): Path to output Excel file
        - pivot_params: Dictionary containing pivot table configuration:
            - filter (str, optional): Column name to use as filter/page field
            - filter_value (Any, optional): filter=filter_value condition
            - row (str, optional): Column name to use as row field
            - row_grouping (tuple, optional): Parameters for grouping row field (start, end, interval)
            - row_edges (array, optional): array of edges to group row field, active when row_grouping is not provided
            - col (str, optional): Column name to use as column field
            - col_grouping (tuple, optional): Parameters for grouping column field (start, end, interval)
            - col_edges (array, optional): array of edges to group column field, active when col_grouping is not provided
            - val (str, optional): Column name to use as values field
            - func (array, optional): List of aggregation function ("SUM", "COUNT", "AVERAGE", "MAX", "MIN")

    Example:    
        datafile = 'data.csv'
        pivot_params = {
            'filter': 'type',
            'filter_value': 'B2C',
            'row': 'buy_price',
            'row_grouping': (-10, 10, 5),
            'col': 'sell_price',
            'col_edges': [0, 1, 3, 4],
            'val': 'commission',
            'func': ['COUNT', 'SUM'],
        }
        create_excel_live_pivot(datafile=datafile, pivot_params=pivot_params)
    """
    SUPPORTED_FUNCS = {
        "SUM": "SUMIFS",
        "COUNT": "COUNTIFS",
        "AVERAGE": "AVERAGEIFS",
        "MAX": "MAXIFS",
        "MIN": "MINIFS"
    }

    df = pd.read_csv(datafile)

    if df.empty:
        print("Data is empty. Aborting.")
        return

    # Extract pivot parameters
    row = pivot_params.get('row')
    row_grouping = pivot_params.get('row_grouping')
    row_edges = pivot_params.get('row_edges')
    col = pivot_params.get('col')
    col_grouping = pivot_params.get('col_grouping')
    col_edges = pivot_params.get('col_edges')
    val = pivot_params.get('val')
    func = pivot_params.get('func')
    filter_col = pivot_params.get('filter')
    filter_value = pivot_params.get('filter_value')

    # Validate row, col, val, filter exist in dataframe
    if row and row not in df.columns:
        print(f"Row column '{row}' not found in dataframe columns")
        row = None
    if col and col not in df.columns:
        print(f"Column '{col}' not found in dataframe columns")
        col = None
    if val and val not in df.columns:
        print(f"Value column '{val}' not found in dataframe columns")
        val = None
    if filter_col and filter_col not in df.columns:
        print(f"Filter column '{filter_col}' not found in dataframe columns")
        filter_col = None

    # Workbook setup
    wb = Workbook()
    ws_pivot = wb.active
    ws_pivot.title = "pivot"
    ws_data = wb.create_sheet("data")
    
    # Write data sheet
    ws_data.append(df.columns.tolist())
    for row_data in df.itertuples(index=False):
        ws_data.append(list(row_data))
    
    # Write pivot sheet

    # Add filter section
    if filter_col:
        filter_col_letter = get_column_letter(df.columns.get_loc(filter_col) + 1)
        ws_pivot.cell(row=1, column=1).value = filter_col
        ws_pivot.cell(row=1, column=1).font = Font(bold=True)
        ws_pivot.cell(row=1, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws_pivot.cell(row=1, column=2).value = filter_value

    # Write name of val column
    if val:
        val_col_letter = get_column_letter(df.columns.get_loc(val) + 1)
        ws_pivot.cell(row=1, column=4).value = val
        ws_pivot.cell(row=1, column=4).font = Font(bold=True)
        ws_pivot.cell(row=1, column=4).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    # Add row headers
    if row:
        row_col_letter = get_column_letter(df.columns.get_loc(row) + 1)
        ws_pivot.cell(row=4, column=1).value = row
        ws_pivot.cell(row=4, column=1).font = Font(bold=True)
        ws_pivot.cell(row=5, column=1).value = "from"
        ws_pivot.cell(row=5, column=1).alignment = Alignment(horizontal='center')
        ws_pivot.cell(row=5, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws_pivot.cell(row=5, column=2).value = "to"
        ws_pivot.cell(row=5, column=2).alignment = Alignment(horizontal='center')
        ws_pivot.cell(row=5, column=2).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        ws_pivot.cell(row=5, column=3).value = "range"
        ws_pivot.cell(row=5, column=3).alignment = Alignment(horizontal='center')
        ws_pivot.cell(row=5, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Add row grouping
        if row_grouping: # override row_edges
            start, end, interval = row_grouping
            row_edges = [start]
            while row_edges[-1] < end:
                row_edges.append(row_edges[-1] + interval)

        if row_edges:
            # Calculate min and max values from the row field and row edges
            min_val = min(df[row].min(), row_edges[0]) - 1e-6
            max_val = max(df[row].max(), row_edges[-1]) + 1e-6
            
            # Add min and max as edges
            row_edges.insert(0, min_val)
            row_edges.append(max_val)

            # adding row headers: from column (column 1), to column (column 2)
            num_rows = len(row_edges) - 1

            # Fill with light gray background
            for i in range(num_rows):
                ws_pivot.cell(row=6+i, column=1).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_pivot.cell(row=6+i, column=2).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # first from-to pair
            ws_pivot.cell(row=6, column=1).value = row_edges[0]
            ws_pivot.cell(row=6, column=1).number_format = '0.00'
            ws_pivot.cell(row=6, column=2).value = row_edges[1]
            ws_pivot.cell(row=6, column=2).number_format = '0.00'
            ws_pivot.cell(row=6, column=2).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            ws_pivot.cell(row=6, column=3).value = f'="<" & ROUND(B6,2)'
            ws_pivot.cell(row=6, column=3).alignment = Alignment(horizontal='center')
            ws_pivot.cell(row=6, column=3).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            # rest of the from-to pairs
            for i in range(1, num_rows):
                ws_pivot.cell(row=6+i, column=1).value = f'=B{6+i-1}'
                ws_pivot.cell(row=6+i, column=1).number_format = '0.00'
                ws_pivot.cell(row=6+i, column=2).value = row_edges[i+1]
                ws_pivot.cell(row=6+i, column=2).number_format = '0.00'
                ws_pivot.cell(row=6+i, column=2).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
                ws_pivot.cell(row=6+i, column=3).value = f'=ROUND(A{6+i},2)&" - "& ROUND(B{6+i},2)'
                ws_pivot.cell(row=6+i, column=3).alignment = Alignment(horizontal='center')
                ws_pivot.cell(row=6+i, column=3).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            ws_pivot.cell(row=6+num_rows-1, column=2).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # last cell

            # Total row
            ws_pivot.cell(row=6+num_rows, column=3).value = "Total"
            ws_pivot.cell(row=6+num_rows, column=3).alignment = Alignment(horizontal='right')
            ws_pivot.cell(row=6+num_rows, column=3).font = Font(bold=True)
            ws_pivot.cell(row=6+num_rows, column=1).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            ws_pivot.cell(row=6+num_rows, column=2).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            ws_pivot.cell(row=6+num_rows, column=3).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Add column headers
    if col:
        col_col_letter = get_column_letter(df.columns.get_loc(col) + 1)
        ws_pivot.cell(row=3, column=4).value = col + " -->"
        ws_pivot.cell(row=3, column=4).font = Font(bold=True)

        # Add column grouping
        if col_grouping: # override col_edges
            start, end, interval = col_grouping
            col_edges = [start]
            while col_edges[-1] < end:
                col_edges.append(col_edges[-1] + interval)

        if col_edges:
            # Calculate min and max values from the column field and column edges
            min_val = min(df[col].min(), col_edges[0]) - 1e-6
            max_val = max(df[col].max(), col_edges[-1]) + 1e-6
            
            # Add min and max as edges
            col_edges.insert(0, min_val)
            col_edges.append(max_val)

            # adding column headers: from row (row 4), to    row (row 5)
            num_cols = len(col_edges) - 1

            # Fill with light gray background
            for i in range(num_cols):
                ws_pivot.cell(row=4, column=4+i).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_pivot.cell(row=5, column=4+i).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # first from-to pair
            ws_pivot.cell(row=4, column=4).value = col_edges[0]
            ws_pivot.cell(row=4, column=4).number_format = '0.00'
            ws_pivot.cell(row=5, column=4).value = col_edges[1]
            ws_pivot.cell(row=5, column=4).number_format = '0.00'
            ws_pivot.cell(row=5, column=4).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

            # rest of the from-to pairs
            for i in range(1, num_cols):
                ws_pivot.cell(row=4, column=4+i).value = f'={get_column_letter(4+i-1)}5'
                ws_pivot.cell(row=4, column=4+i).number_format = '0.00'
                ws_pivot.cell(row=5, column=4+i).value = col_edges[i+1]
                ws_pivot.cell(row=5, column=4+i).number_format = '0.00'
                ws_pivot.cell(row=5, column=4+i).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            ws_pivot.cell(row=5, column=4+num_cols-1).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid") # last cell

    # Table Type 1: col headers are provided, then admit only the first function in func, and each cell contains the formula with conditions on both row and column
    # Table Type 2: col headers are not provided, then admit all functions in func, and each cell contains the formula with conditions on row only
    if col: # Table Type 1
        func_name = func[0] # only the first function is supported
        if func_name not in SUPPORTED_FUNCS:
            print(f"Unsupported aggregation function: {func_name}. Defaulting to COUNT.")
            func_name = "COUNT"

        # Write name of the func_name to E1
        ws_pivot.cell(row=1, column=5).value = func_name
        ws_pivot.cell(row=1, column=5).font = Font(bold=True)

        for i in range(num_rows):
            col_idx = 3
            for j in range(num_cols):
                col_idx += 1
                conditions = []
                if filter_col:
                    conditions.append(f'data!${filter_col_letter}:${filter_col_letter}, $B$1')

                if row_col_letter:
                    conditions.append(f'data!${row_col_letter}:${row_col_letter}, ">=" & $A{6+i}')
                    conditions.append(f'data!${row_col_letter}:${row_col_letter}, "<" & $B{6+i}')

                if col_col_letter:
                    conditions.append(f'data!${col_col_letter}:${col_col_letter}, ">=" & {get_column_letter(col_idx)}$4')
                    conditions.append(f'data!${col_col_letter}:${col_col_letter}, "<" & {get_column_letter(col_idx)}$5')
                    
                if func_name == "COUNT":
                    formula = f'=COUNTIFS({", ".join(conditions)})'
                else:
                    agg_func = SUPPORTED_FUNCS[func_name]
                    formula = f'={agg_func}(data!{val_col_letter}:{val_col_letter},{", ".join(conditions)})'
                    ws_pivot.cell(row=6+i, column=col_idx).number_format = '#,##0;(#,##0)'

                ws_pivot.cell(row=6+i, column=col_idx).value = formula

        # Add a Total column to the right
        ws_pivot.cell(row=5, column=num_cols+4).value = "Total"
        ws_pivot.cell(row=5, column=num_cols+4).alignment = Alignment(horizontal='center')
        ws_pivot.cell(row=5, column=num_cols+4).font = Font(bold=True)
        ws_pivot.cell(row=5, column=num_cols+4).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        for i in range(6, num_rows+7):
            ws_pivot.cell(row=i, column=num_cols+4).value = f'=SUM({get_column_letter(4)}{i}:{get_column_letter(num_cols+3)}{i})'
            ws_pivot.cell(row=i, column=num_cols+4).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            if func_name != "COUNT":
                ws_pivot.cell(row=i, column=num_cols+4).number_format = '#,##0;(#,##0)'

        # Total row
        for col_idx in range(4, num_cols+5):
            ws_pivot.cell(row=6+num_rows, column=col_idx).value = f'=SUM({get_column_letter(col_idx)}6:{get_column_letter(col_idx)}{5+num_rows})'
            ws_pivot.cell(row=6+num_rows, column=col_idx).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
            if func_name != "COUNT":
                ws_pivot.cell(row=6+num_rows, column=col_idx).number_format = '#,##0;(#,##0)'
            
        # Adjust column width to fit content
        for col_idx in range(4, num_cols+5):
            if func_name != "COUNT":
                ws_pivot.column_dimensions[get_column_letter(col_idx)].width = 12
                if col_idx == num_cols+4: # different width for total column
                    ws_pivot.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Draw lines
        for i in range(1, num_cols+5):
            ws_pivot.cell(row=5, column=i).border = Border(bottom=Side(style='thin'))
            ws_pivot.cell(row=6+num_rows, column=i).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        # set active cell
        ws_pivot.sheet_view.selection[0].activeCell = f'{get_column_letter(4+num_cols)}{6+num_rows}'
        ws_pivot.sheet_view.selection[0].sqref = f'{get_column_letter(4+num_cols)}{6+num_rows}'

    else: # Table Type 2
        if func and row:
            col_idx = 3
            for func_name in func:
                col_idx += 1

                if func_name not in SUPPORTED_FUNCS:
                    print(f"Unsupported aggregation function: {func_name}. Defaulting to COUNT.")
                    func_name = "COUNT"
                
                ws_pivot.cell(row=5, column=col_idx).value = func_name
                ws_pivot.cell(row=5, column=col_idx).alignment = Alignment(horizontal='center')
                ws_pivot.cell(row=5, column=col_idx).font = Font(bold=True)
                ws_pivot.cell(row=5, column=col_idx).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
                # Add function formula
                for i in range(num_rows):
                    conditions = []
                    if filter_col:
                        conditions.append(f'data!${filter_col_letter}:${filter_col_letter}, $B$1')

                    if row_col_letter:
                        conditions.append(f'data!${row_col_letter}:${row_col_letter}, ">=" & $A{6+i}')
                        conditions.append(f'data!${row_col_letter}:${row_col_letter}, "<" & $B{6+i}')
                                        
                    if func_name == "COUNT":
                        formula = f'=COUNTIFS({", ".join(conditions)})'
                    else:
                        agg_func = SUPPORTED_FUNCS[func_name]
                        formula = f'={agg_func}(data!{val_col_letter}:{val_col_letter}, {", ".join(conditions)})'
                        ws_pivot.cell(row=6+i, column=col_idx).number_format = '#,##0;(#,##0)'

                    ws_pivot.cell(row=6+i, column=col_idx).value = formula

                # Total row
                ws_pivot.cell(row=6+num_rows, column=col_idx).value = f'=SUM({get_column_letter(col_idx)}6:{get_column_letter(col_idx)}{5+num_rows})'
                ws_pivot.cell(row=6+num_rows, column=col_idx).fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                if func_name != "COUNT":
                    ws_pivot.cell(row=6+num_rows, column=col_idx).number_format = '#,##0;(#,##0)'

                # Adjust column width to fit content
                if func_name != "COUNT":
                    ws_pivot.column_dimensions[get_column_letter(col_idx)].width = 12

            # set active cell
            ws_pivot.sheet_view.selection[0].activeCell = f'{get_column_letter(col_idx)}{6+num_rows}'
            ws_pivot.sheet_view.selection[0].sqref = f'{get_column_letter(col_idx)}{6+num_rows}'
                                                    
    wb.save(outputfile)
    print(f"Live pivot Excel created: {outputfile}")


def create_excel_pivot_native(datafile=None, outputfile=None, pivot_params=None):
    """
    Create live Excel pivot table from a CSV file, using native Excel pivot table.

    Require: Windows with Excel installed
    """
    
    datafile = datafile or 'output/trade_exit_log.csv'
    outputfile = outputfile or 'output/excel_pivot.xlsx'

    EXCEL_ORIENTATIONS = {
        'ROW': 1,      # xlRowField
        'COLUMN': 2,   # xlColumnField
        'PAGE': 3,     # xlPageField
        'DATA': 4      # xlDataField
    }

    EXCEL_FUNCTIONS = {
        "SUM": -4157,      # xlSum
        "COUNT": -4112,    # xlCount
        "AVERAGE": -4106,  # xlAverage
        "MAX": -4102,      # xlMax
        "MIN": -4107,      # xlMin
        "PRODUCT": -4149,  # xlProduct
    }

    @contextmanager
    def excel_session():
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        pid = win32process.GetWindowThreadProcessId(excel.Hwnd)[1]
        try:
            yield excel
        finally:
            excel.Quit()
            if psutil.pid_exists(pid):
                psutil.Process(pid).kill()

    data_sheet_name = "data"
    pivot_sheet_name = "pivot"

    df = pd.read_csv(datafile)
    
    if df.empty:
        print("Data is empty. Pivot table not created.")
        return
    
    # Export df to Excel outputfile
    outputfile = os.path.abspath('excel_pivot.xlsx' if outputfile is None else outputfile)
    if os.path.exists(outputfile):
        try:
            with open(outputfile, 'a+b') as f:
                pass
        except Exception as e:
            print(f"\nError accessing file: {str(e)}")
            return
    try:
        df.to_excel(outputfile, sheet_name=data_sheet_name, index=False)
    except Exception as e:
        print(f"\nError writing to file: {str(e)}")
        return
    
    # Validate input parameters
    valid_field_names = df.columns.tolist()
    filter_field = pivot_params.get('filter') if pivot_params.get('filter') in valid_field_names else None
    filter_value = pivot_params.get('filter_value') if (filter_field is not None) else None
    row = pivot_params.get('row') if pivot_params.get('row') in valid_field_names else None
    row_grouping = pivot_params.get('row_grouping')
    col = pivot_params.get('col') if pivot_params.get('col') in valid_field_names else None
    col_grouping = pivot_params.get('col_grouping')
    val = pivot_params.get('val') if pivot_params.get('val') in valid_field_names else None
    try:
        with excel_session() as excel:
            wb = None
            try:
                wb = excel.Workbooks.Open(outputfile)
                ws = wb.Worksheets(data_sheet_name)
                pivot_ws = wb.Worksheets.Add()
                pivot_ws.Name = pivot_sheet_name
                last_row, last_col = ws.UsedRange.Rows.Count, ws.UsedRange.Columns.Count
                pcache = wb.PivotCaches().Create(SourceType=1, SourceData=f"'{data_sheet_name}'!R1C1:R{last_row}C{last_col}")
                ptable = pcache.CreatePivotTable(TableDestination=pivot_ws.Cells(1, 1), TableName="PivotTable1")
                pivot_ws.Columns.ColumnWidth = 15

                if filter_field:
                    ptable.PivotFields(filter_field).Orientation = EXCEL_ORIENTATIONS['PAGE']
                    if filter_value:
                        ptable.PivotFields(filter_field).CurrentPage = str(filter_value) # Excel expects a string
                
                if row:
                    ptable.PivotFields(row).Orientation = EXCEL_ORIENTATIONS['ROW']    
                    if row_grouping:
                        f = ptable.RowFields.Item(1)
                        pivot_ws.Cells(f.LabelRange.Row + 1, f.LabelRange.Column).Select()
                        excel.Selection.Group(Start=row_grouping[0], End=row_grouping[1], By=row_grouping[2])

                if col:
                    ptable.PivotFields(col).Orientation = EXCEL_ORIENTATIONS['COLUMN']
                    if col_grouping:
                        f = ptable.ColumnFields.Item(1)
                        pivot_ws.Cells(f.LabelRange.Row + 1, f.LabelRange.Column).Select()
                        excel.Selection.Group(Start=col_grouping[0], End=col_grouping[1], By=col_grouping[2])

                if val:
                    # Table Type 1: col headers are provided, then admit only the first function in func
                    # Table Type 2: col headers are not provided, then admit all functions in func
                    if col:
                        func = pivot_params.get('func')[:1]
                    else:
                        func = pivot_params.get('func')
                                        
                    for func_name in func:
                        if func_name not in EXCEL_FUNCTIONS.keys():
                            print(f"Unsupported aggregation function: {func_name}. Defaulting to COUNT.")
                            func_name = "COUNT"
                        data_field = ptable.AddDataField(ptable.PivotFields(val), f"{func_name} of {val}", EXCEL_FUNCTIONS.get(func_name.upper()))
                        
                        # Pre-format for numeric value fields
                        if isinstance(df[val].iloc[0], (int, float)):
                            data_field.NumberFormat = "#,##0;(#,##0)"

                    # hard-coded pnlcomm_win to pivot table Type 2 for now. TODO: get this from the input parameters
                    if not col and True:
                        val_tm = "pnlcomm_win"
                        func_name_tm = "AVERAGE"
                        data_field = ptable.AddDataField(ptable.PivotFields(val_tm), f"{func_name_tm} of {val_tm}", EXCEL_FUNCTIONS.get(func_name_tm.upper()))
                        data_field.NumberFormat = "0.0%"

                ptable.RowGrand = True
                ptable.ColumnGrand = True
                pivot_ws.Columns.ColumnWidth = 12
                pivot_ws.Rows.AutoFit()

                wb.Save()
                wb.Close()
                print(f"Pivot table successfully created and saved to: {outputfile}")
            except Exception as e:
                if wb:
                    try:
                        wb.Close(SaveChanges=False)
                    except:
                        pass
                print(f"\nError: {str(e)}")
                return
    except Exception as e:
        print(f"\nError: {str(e)}")
        return

# test code
if __name__ == "__main__":
    datafile = 'output/trade_exit_log.csv'
    outputfile = 'output/excel_pivot.xlsx'
    pivot_params = {
        'row': 'adx',
        'val': 'pnl',
        'func': ['COUNT', 'SUM', 'AVERAGE']
    }
    create_excel_pivot_native(datafile=datafile, outputfile=outputfile, pivot_params=pivot_params)
